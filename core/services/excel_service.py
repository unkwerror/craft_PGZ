#core/services/excel_service.py
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.tender import Tender


class ExcelService:
    def __init__(self, base_output_dir="results"):
        self.base_output_dir = Path(base_output_dir)

    def save_tender(self, tender: Tender) -> Path:
        filepath = self.base_output_dir / tender.reg_number / f"{tender.reg_number}.xlsx"
        filepath.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Данные закупки"

        headers = ["Параметр", "Значение"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        rows = [
            ["Номер заявки", tender.reg_number],
            ["Название закупки", tender.title],
            ["Сумма контракта", tender.price],
            ["Дата окончания подачи", tender.end_date],
        ]
        for row_num, row in enumerate(rows, 2):
            for col_num, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")

        wb.save(filepath)
        return filepath
