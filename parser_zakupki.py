import os
import time
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urljoin


class ZakupkiParser:
    def __init__(self, base_output_dir="results"):
        self.driver = self._init_driver()
        self.base_url = "https://zakupki.gov.ru/epz/order/notice/ea20/view/common-info.html?regNumber="
        self.base_output_dir = base_output_dir
        # Создаем базовую директорию если ее нет
        Path(base_output_dir).mkdir(exist_ok=True)

    def _init_driver(self) -> webdriver.Chrome:
        options = Options()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        user_agent = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/140.0.0.0 Safari/537.36"
        )
        options.add_argument(f'user-agent={user_agent}')
        return webdriver.Chrome(options=options)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.driver.quit()

    def create_tender_folder(self, reg_number: str) -> Path:
        """Создает папку для результатов парсинга с номером закупки"""
        tender_path = Path(self.base_output_dir) / reg_number
        tender_path.mkdir(exist_ok=True)
        return tender_path

    def load_page(self, reg_number: str):
        url = self.base_url + reg_number
        self.driver.get(url)
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".cardMainInfo"))
        )
        time.sleep(1)

    def click_documents_tab(self):
        """Кликает на вкладку 'Документы' с обходом проблем с перекрытием"""
        try:
            # Ждем появления элемента
            documents_tab = WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[1]/div[3]/div/a[2]"))
            )
            
            # Прокручиваем к элементу
            self.driver.execute_script("arguments[0].scrollIntoView(true);", documents_tab)
            time.sleep(1)
            
            # Пробуем кликнуть через JavaScript
            self.driver.execute_script("arguments[0].click();", documents_tab)
            
            # Ждем загрузки документов
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CLASS_NAME, "blockFilesTabDocs"))
            )
            time.sleep(1)
            return True
        except Exception as e:
            print(f"Ошибка при переходе на вкладку 'Документы': {e}")
            
            # Пробуем альтернативный метод - клик по ссылке через URL
            try:
                reg_number = self.driver.current_url.split("regNumber=")[-1]
                documents_url = f"https://zakupki.gov.ru/epz/order/notice/ea20/view/documents.html?regNumber={reg_number}"
                self.driver.get(documents_url)
                
                # Ждем загрузки документов
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "blockFilesTabDocs"))
                )
                time.sleep(1)
                return True
            except Exception as e2:
                print(f"Ошибка при прямом переходе на страницу документов: {e2}")
                return False

    def expand_all_documents(self):
        """Кликает на 'Показать больше' чтобы отобразить все документы"""
        try:
            show_more_button = self.driver.find_element(
                By.XPATH, "//a[contains(text(), 'Показать больше')]"
            )
            show_more_button.click()
            time.sleep(1)
            return True
        except:
            # Если кнопки нет, значит все документы уже отображены
            return False

    def parse_documents(self):
        """Парсит список документов на странице"""
        documents = []
        
        # Попробуем развернуть все документы
        self.expand_all_documents()
        
        try:
            # Более надежный способ поиска документов
            doc_containers = self.driver.find_elements(
                By.CSS_SELECTOR, ".blockFilesTabDocs .attachment"
            )
            
            for container in doc_containers:
                try:
                    # Ищем ссылку внутри контейнера
                    link_element = container.find_element(
                        By.CSS_SELECTOR, ".section__value a, a[href*='download']"
                    )
                    name = link_element.text.strip()
                    url = link_element.get_attribute("href")
                    
                    # Получаем оригинальное имя файла
                    original_filename = link_element.get_attribute("title")
                    if not original_filename:
                        # Пробуем получить из атрибута data-filename если есть
                        original_filename = link_element.get_attribute("data-filename")
                        if not original_filename:
                            # Используем текст ссылки как запасной вариант
                            original_filename = name
                    
                    documents.append({
                        "name": name,
                        "url": url,
                        "original_filename": original_filename
                    })
                except Exception as e:
                    print(f"Ошибка при парсинге документа в контейнере: {e}")
                    continue
                    
        except Exception as e:
            print(f"Ошибка при поиске контейнеров документов: {e}")
            
            # Альтернативный метод поиска документов
            try:
                doc_links = self.driver.find_elements(
                    By.CSS_SELECTOR, ".blockFilesTabDocs a[href*='download']"
                )
                
                for link in doc_links:
                    try:
                        name = link.text.strip()
                        url = link.get_attribute("href")
                        original_filename = link.get_attribute("title") or name
                        
                        documents.append({
                            "name": name,
                            "url": url,
                            "original_filename": original_filename
                        })
                    except Exception as e:
                        print(f"Ошибка при парсинге документа по ссылке: {e}")
                        continue
            except Exception as e2:
                print(f"Ошибка при альтернативном поиске документов: {e2}")
        
        return documents
    def download_document(self, url, filepath):
        """Скачивает документ по URL и сохраняет по указанному пути"""
        try:
            # Получаем cookies из браузера
            cookies = self.driver.get_cookies()
            session = requests.Session()
            
            # Добавляем cookies в сессию
            for cookie in cookies:
                session.cookies.set(cookie['name'], cookie['value'])
            
            # Устанавливаем заголовки как в браузере
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
                'Referer': self.driver.current_url,
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br',
            }
            
            # Скачиваем файл
            response = session.get(url, headers=headers, stream=True, timeout=30)
            response.raise_for_status()
            
            # Определяем расширение файла из Content-Type если нужно
            content_type = response.headers.get('Content-Type', '')
            if 'application/pdf' in content_type and not str(filepath).lower().endswith('.pdf'):
                filepath = Path(str(filepath) + '.pdf')
            elif 'application/msword' in content_type and not str(filepath).lower().endswith('.doc'):
                filepath = Path(str(filepath) + '.doc')
            elif 'vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type and not str(filepath).lower().endswith('.docx'):
                filepath = Path(str(filepath) + '.docx')
            elif 'application/vnd.ms-excel' in content_type and not str(filepath).lower().endswith('.xls'):
                filepath = Path(str(filepath) + '.xls')
            elif 'vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type and not str(filepath).lower().endswith('.xlsx'):
                filepath = Path(str(filepath) + '.xlsx')
            
            # Сохраняем файл
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            return True
        except Exception as e:
            print(f"Ошибка при скачивании файла {url}: {e}")
            return False

    def download_all_documents(self, reg_number):
        """Скачивает все документы закупки"""
        documents = self.parse_documents()
        tender_path = self.create_tender_folder(reg_number)
        docs_path = tender_path / "documents"
        docs_path.mkdir(exist_ok=True)
        
        downloaded_files = []
        
        for doc in documents:
            try:
                # Создаем безопасное имя файла
                safe_filename = "".join(
                    c for c in doc['original_filename'] 
                    if c.isalpha() or c.isdigit() or c in (' ', '-', '_', '.')
                ).rstrip()
                
                filepath = docs_path / safe_filename
                
                # Скачиваем документ
                if self.download_document(doc['url'], filepath):
                    downloaded_files.append({
                        "name": doc['name'],
                        "filename": safe_filename,
                        "path": str(filepath)
                    })
                    print(f"Скачан документ: {safe_filename}")
                
                # Небольшая задержка между запросами
                time.sleep(0.5)
                
            except Exception as e:
                print(f"Ошибка при обработке документа {doc['name']}: {e}")
        
        return downloaded_files

    def save_html(self, reg_number: str, filename="page.html"):
        """Сохраняет HTML в папку с номером закупки"""
        tender_path = self.create_tender_folder(reg_number)
        filepath = tender_path / filename
        html = self.driver.page_source
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

    def save_to_excel(self, data: dict):
        """Сохраняет данные в Excel файл в папке закупки"""
        tender_path = self.create_tender_folder(data["reg_number"])
        filepath = tender_path / f"{data['reg_number']}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Данные закупки"

        # Заголовки
        headers = ["Параметр", "Значение"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

        # Данные
        rows = [
            ["Номер заявки", data.get("reg_number", "")],
            ["Название закупки", data.get("title", "")],
            ["Сумма контракта", data.get("price", "")],
            ["Дата окончания подачи", data.get("end_date", "")]
        ]

        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

        wb.save(filepath)

    def parse_tender_card(self) -> dict:
        """Парсит карточку закупки в соответствии с ТЗ"""
        data = {}

        def safe_xpath(xpath: str):
            try:
                return self.driver.find_element(By.XPATH, xpath).text.strip()
            except:
                return None

        data["reg_number"] = self.driver.current_url.split("regNumber=")[-1]
        data["title"] = safe_xpath('/html/body/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[2]/div[1]/span[2]')
        data["price"] = safe_xpath('/html/body/div[2]/div/div[1]/div[2]/div[2]/div[2]/div[1]/span[2]')
        data["end_date"] = safe_xpath('/html/body/div[2]/div/div[1]/div[2]/div[2]/div[2]/div[2]/div[2]/span[2]')

        return data


if __name__ == "__main__":
    reg_number = "0162200011825003267"
    
    with ZakupkiParser() as parser:
        # Загружаем основную страницу
        parser.load_page(reg_number)
        info = parser.parse_tender_card()
        print(info)
        
        # Сохраняем основную информацию
        parser.save_to_excel(info)
        parser.save_html(reg_number)
        
        # Переходим на вкладку документов и скачиваем их
        if parser.click_documents_tab():
            # Сохраняем HTML страницы с документами
            parser.save_html(reg_number, "documents_page.html")
            
            # Скачиваем все документы
            downloaded = parser.download_all_documents(reg_number)
            print(f"Скачано документов: {len(downloaded)}")